import socket
import ssl
import datetime
import dns.resolver
import requests
import openpyxl
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed
import warnings
import re
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse

# Suppress InsecureRequestWarning for self-signed certs if verify=False is used
from requests.packages.urllib3.exceptions import InsecureRequestWarning
warnings.simplefilter('ignore', InsecureRequestWarning)

COMMON_PORTS = {
    21: "FTP", 22: "SSH", 23: "Telnet", 25: "SMTP", 53: "DNS",
    80: "HTTP", 110: "POP3", 143: "IMAP", 443: "HTTPS", 445: "SMB",
    993: "IMAPS", 995: "POP3S", 3306: "MySQL", 3389: "RDP",
    5432: "PostgreSQL", 5900: "VNC", 8080: "HTTP Alt", 8443: "HTTPS Alt"
}

# Basic regex patterns for finding secrets - CAUTION: HIGH FALSE POSITIVES LIKELY
# These are illustrative and would need significant refinement for real-world use.
SECRET_PATTERNS = {
    "API_KEY": re.compile(r"""
        (?:["']?)                                                                              # Optional starting quote (non-capturing)
        (?:api_key|apikey|api_token|access_key|access_token|client_secret|secret_key|auth_token) # Common key names (non-capturing)
        (?:["']?)                                                                              # Optional ending quote for the key name (non-capturing)
        \s*[:=]\s*                                                                             # Separator (colon or equals)
        (?:["']?)                                                                              # Optional starting quote for the value (non-capturing)
        ([a-zA-Z0-9_/\-+]{16,64})                                                              # The key value itself (CAPTURING GROUP 1)
        (?:["']?)                                                                              # Optional ending quote for the value (non-capturing)
    """, re.IGNORECASE | re.VERBOSE),
    "AWS_ACCESS_KEY_ID": re.compile(r"AKIA[0-9A-Z]{16}"),
    "AWS_SECRET_KEY": re.compile(r"(?<![A-Z0-9])[A-Za-z0-9/+=]{40}(?![A-Z0-9])"), # More generic, prone to FPs
    "GENERIC_PASSWORD_ASSIGN": re.compile(r"""
        (?:["']?)                     # Optional starting quote (non-capturing)
        (?:password|passwd|pwd)       # Common password names (non-capturing)
        (?:["']?)                     # Optional ending quote for the key name (non-capturing)
        \s*[:=]\s*                    # Separator
        (?:["']?)                     # Optional starting quote for the value (non-capturing)
        ([^"'\s]{6,})                 # Password value (at least 6 non-space, non-quote chars) (CAPTURING GROUP 1)
        (?:["']?)                     # Optional ending quote for the value (non-capturing)
    """, re.IGNORECASE | re.VERBOSE),
    "BEARER_TOKEN": re.compile(r"Bearer\s+([a-zA-Z0-9_.\-]{20,})", re.IGNORECASE), # Note: re.VERBOSE not used here, so # comments are not allowed inside this raw string
    "PRIVATE_KEY_HEADER": re.compile(r"-----BEGIN ((RSA|OPENSSH|EC|PGP) )?PRIVATE KEY-----"), # Note: re.VERBOSE not used here
}

# --- Enumeration Functions ---

def get_dns_records(domain):
    """Gets various DNS records for a domain."""
    records = {'A': [], 'AAAA': [], 'CNAME': [], 'MX': [], 'TXT': [], 'NS': []}
    record_types = ['A', 'AAAA', 'CNAME', 'MX', 'TXT', 'NS']
    
    for r_type in record_types:
        try:
            answers = dns.resolver.resolve(domain, r_type)
            for rdata in answers:
                if r_type == 'MX':
                    records[r_type].append(f"{rdata.preference} {rdata.exchange.to_text()}")
                elif r_type == 'TXT':
                    records[r_type].extend([txt.decode('utf-8', 'ignore') for txt in rdata.strings])
                else:
                    records[r_type].append(rdata.to_text())
        except dns.resolver.NXDOMAIN:
            records[r_type].append("NXDOMAIN") 
            if r_type == 'A': 
                break 
        except dns.resolver.NoAnswer:
            records[r_type].append("NoAnswer")
        except dns.resolver.Timeout:
            records[r_type].append("Timeout")
        except Exception: # Catching generic exception for DNS part
            records[r_type].append("DNS Resolution Error")
    return records

def check_ports(domain_or_ip, ports_to_check=None):
    if ports_to_check is None:
        ports_to_check = COMMON_PORTS.keys()
    open_ports = []
    target_ip = ""
    try:
        target_ip = socket.gethostbyname(domain_or_ip)
    except socket.gaierror:
        return {"ip_resolved": "Failed to resolve", "open_ports_details": "N/A"}

    for port in ports_to_check:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(0.5)
        result = sock.connect_ex((target_ip, port))
        if result == 0:
            port_name = COMMON_PORTS.get(port, "Unknown")
            open_ports.append(f"{port}/{port_name}")
        sock.close()
    return {"ip_resolved": target_ip, "open_ports_details": ", ".join(open_ports) if open_ports else "None found"}

def find_secrets_in_content(content_url, text_content):
    """Scans text content for predefined secret patterns."""
    found_secrets_list = []
    lines = text_content.splitlines()
    for i, line in enumerate(lines):
        if len(line) > 5000: continue # Skip extremely long lines to avoid regex DoS

        for secret_type, pattern in SECRET_PATTERNS.items():
            matches = pattern.finditer(line)
            for match in matches:
                # Try to get the matched group that contains the secret value itself
                secret_value = match.group(1) if len(match.groups()) > 0 else match.group(0)
                # Basic obfuscation for display, show only part of it
                obfuscated_value = secret_value[:4] + "****" + secret_value[-4:] if len(secret_value) > 8 else secret_value
                
                found_secrets_list.append(
                    f"Type: {secret_type}, "
                    f"Value (obfuscated): {obfuscated_value}, "
                    f"Line: {i+1}, "
                    f"File: {content_url}"
                )
    return found_secrets_list


def get_http_info_and_js_secrets(domain):
    """Gets HTTP/HTTPS headers, basic info, and scans linked JS files for secrets."""
    results = {
        'http_status': 'N/A', 'http_headers': {}, 'https_status': 'N/A', 
        'https_headers': {}, 'directory_listing_http': 'N/A', 
        'directory_listing_https': 'N/A', 'final_url_http': 'N/A',
        'final_url_https': 'N/A', 'security_headers_http': {}, 'security_headers_https': {},
        'js_secrets_found': []
    }
    
    protocols = {'http': 80, 'https': 443}
    security_header_names = [
        "Strict-Transport-Security", "X-Frame-Options", "X-Content-Type-Options",
        "Content-Security-Policy", "Referrer-Policy", "Permissions-Policy"
    ]
    all_found_secrets = []

    for proto in protocols:
        base_url = f"{proto}://{domain}"
        page_content = ""
        try:
            response = requests.get(base_url, timeout=7, allow_redirects=True, verify=False, stream=False)
            status_key = f'{proto}_status'
            headers_key = f'{proto}_headers'
            dir_list_key = f'directory_listing_{proto}'
            final_url_key = f'final_url_{proto}'
            sec_headers_key = f'security_headers_{proto}'

            results[status_key] = response.status_code
            results[final_url_key] = response.url 
            page_content = response.text # Store page content for JS parsing

            relevant_headers = {}
            for h_name in ["Server", "X-Powered-By", "Content-Type", "Set-Cookie", "Via", "X-AspNet-Version"]:
                if h_name in response.headers:
                    relevant_headers[h_name] = response.headers[h_name]
            results[headers_key] = relevant_headers
            
            if "Index of /" in page_content or "Parent Directory" in page_content:
                results[dir_list_key] = "Potentially Exposed"
            else:
                results[dir_list_key] = "Not Detected"
            
            sec_headers_found = {}
            for sh_name in security_header_names:
                sec_headers_found[sh_name] = response.headers.get(sh_name, "Not Found")
            results[sec_headers_key] = sec_headers_found

            # --- Find and scan JavaScript files ---
            if response.status_code == 200 and 'text/html' in response.headers.get('Content-Type', '').lower():
                soup = BeautifulSoup(page_content, 'html.parser')
                script_tags = soup.find_all('script', src=True)
                js_urls_to_scan = set()

                for tag in script_tags:
                    js_url = tag['src']
                    # Construct absolute URL if relative
                    js_absolute_url = urljoin(response.url, js_url) 
                    # Optional: Filter to only scan JS from the same domain or subdomains
                    # parsed_js_url = urlparse(js_absolute_url)
                    # parsed_base_url = urlparse(response.url)
                    # if parsed_js_url.netloc == parsed_base_url.netloc or parsed_js_url.netloc.endswith("." + parsed_base_url.netloc):
                    js_urls_to_scan.add(js_absolute_url)

                for js_file_url in js_urls_to_scan:
                    try:
                        print(f"    Fetching JS: {js_file_url[:100]}...") # Print shortened URL
                        js_response = requests.get(js_file_url, timeout=5, verify=False)
                        if js_response.status_code == 200:
                            js_content = js_response.text
                            secrets_in_js = find_secrets_in_content(js_file_url, js_content)
                            if secrets_in_js:
                                all_found_secrets.extend(secrets_in_js)
                        else:
                            all_found_secrets.append(f"JS Fetch Error: {js_file_url} (Status: {js_response.status_code})")
                    except requests.exceptions.RequestException as js_e:
                        all_found_secrets.append(f"JS Fetch Error: {js_file_url} ({type(js_e).__name__})")
                    except Exception as e_js_scan:
                         all_found_secrets.append(f"JS Scan Error: {js_file_url} ({type(e_js_scan).__name__})")


        except requests.exceptions.RequestException as e:
            results[f'{proto}_status'] = f"Error: {type(e).__name__}"
            results[f'directory_listing_{proto}'] = "N/A (Request Error)"
        except Exception as e_http: # Catch broader exceptions during HTTP processing
            results[f'{proto}_status'] = f"General Error: {type(e_http).__name__}"

    results['js_secrets_found'] = all_found_secrets if all_found_secrets else ["No secrets found in linked JS files."]
    return results

def check_ssl_certificate(hostname, port=443):
    context = ssl.create_default_context()
    cert_info = {'issuer': 'N/A', 'expiry_date': 'N/A', 'days_left': 'N/A', 'error': None, 'status': 'N/A'}
    try:
        ip_address = socket.gethostbyname(hostname)
        with socket.create_connection((ip_address, port), timeout=3) as sock:
            with context.wrap_socket(sock, server_hostname=hostname) as ssock:
                cert = ssock.getpeercert() # Removed semicolon
                issuer_parts = [item[0] for item in cert.get('issuer', [])]
                cert_info['issuer'] = ", ".join([f"{part[0]}={part[1]}" for part in issuer_parts])
                if 'notAfter' in cert:
                    expiry_str = cert['notAfter']
                    expiry_date = datetime.datetime.strptime(expiry_str, "%b %d %H:%M:%S %Y %Z")
                    cert_info['expiry_date'] = expiry_date.strftime("%Y-%m-%d %H:%M:%S %Z")
                    days_left = (expiry_date - datetime.datetime.utcnow()).days
                    cert_info['days_left'] = days_left
                    if days_left < 0: cert_info['status'] = "EXPIRED"
                    elif days_left < 30: cert_info['status'] = "Expires Soon (<=30 days)"
                    else: cert_info['status'] = "Valid"
                else:
                    cert_info['status'] = "No Expiry Date Info"
    except socket.gaierror: cert_info['error'] = "Hostname resolution failed"
    except socket.timeout: cert_info['error'] = f"Connection to {hostname}:{port} timed out"
    except ssl.SSLError as e: cert_info['error'] = f"SSL Error: {str(e)}"
    except ConnectionRefusedError: cert_info['error'] = f"Connection to {hostname}:{port} refused"
    except Exception as e: cert_info['error'] = f"SSL Cert An unexpected error: {type(e).__name__}"
    return cert_info

# --- Main Processing and Output ---

def process_target(target):
    print(f"Processing: {target}...")
    data = {'target': target}
    
    data.update(get_dns_records(target))
    
    resolved_ip = data.get('A', [None])[0] if isinstance(data.get('A'), list) else None
    scan_host = resolved_ip if resolved_ip and resolved_ip not in ["NXDOMAIN", "NoAnswer", "Timeout", "DNS Resolution Error"] else target
    
    if scan_host and not any(err_msg in str(scan_host) for err_msg in ["NXDOMAIN", "NoAnswer", "Timeout", "DNS Resolution Error"]):
        port_scan_results = check_ports(scan_host)
        data['ip_resolved_for_scan'] = port_scan_results.get('ip_resolved', 'N/A')
        data['open_ports'] = port_scan_results.get('open_ports_details', 'N/A')
    else:
        data['ip_resolved_for_scan'] = 'N/A (No valid A record or DNS error)'
        data['open_ports'] = 'N/A (No valid A record or DNS error)'

    http_js_data = get_http_info_and_js_secrets(target) # Changed function call
    data.update(http_js_data)

    is_https_likely = (isinstance(http_js_data.get('https_status'), int) and 200 <= http_js_data.get('https_status') < 400) or \
                      ("443/HTTPS" in data.get('open_ports', ''))
                      
    if is_https_likely:
        ssl_data = check_ssl_certificate(target)
        data['ssl_issuer'] = ssl_data.get('issuer')
        data['ssl_expiry_date'] = ssl_data.get('expiry_date')
        data['ssl_days_left'] = ssl_data.get('days_left')
        data['ssl_status'] = ssl_data.get('status')
        data['ssl_error'] = ssl_data.get('error')
    else:
        data.update({'ssl_issuer': 'N/A (HTTPS not detected)', 'ssl_expiry_date': 'N/A', 
                     'ssl_days_left': 'N/A', 'ssl_status': 'N/A (HTTPS not detected)', 'ssl_error': None})

    issues = []
    if data.get('directory_listing_http') == "Potentially Exposed": issues.append("HTTP Directory Listing")
    if data.get('directory_listing_https') == "Potentially Exposed": issues.append("HTTPS Directory Listing")
    if data.get('ssl_status') == "EXPIRED": issues.append("SSL Certificate Expired")
    if data.get('ssl_status') == "Expires Soon (<=30 days)": issues.append("SSL Certificate Expires Soon")
    if data.get('ssl_error'): issues.append(f"SSL Connection Error: {data['ssl_error']}")
    
    for proto in ['http', 'https']:
        sec_headers = data.get(f'security_headers_{proto}', {})
        for h_name, h_value in sec_headers.items():
            if h_value == "Not Found": issues.append(f"Missing Security Header ({proto.upper()}): {h_name}")
    
    js_secrets = data.get('js_secrets_found', [])
    if js_secrets and js_secrets != ["No secrets found in linked JS files."]:
        # Count actual findings, not error messages
        actual_secret_findings = [s for s in js_secrets if not s.startswith("JS Fetch Error") and not s.startswith("JS Scan Error")]
        if actual_secret_findings:
            issues.append(f"Potential Secrets Found in JS ({len(actual_secret_findings)} instances)")


    data['potential_issues'] = ", ".join(issues) if issues else "None obvious from basic checks"
    
    return data

def write_to_excel(data_list, filename="domain_assessment_results.xlsx"):
    if not data_list:
        print("No data to write.")
        return

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Assessment Results"

    headers = [
        'Target', 'IP Resolved (Scan)', 'A Records', 'AAAA Records', 'CNAME Records', 'MX Records', 'NS Records', 'TXT Records',
        'Open Ports', 
        'HTTP Status', 'HTTP Final URL', 'HTTP Server', 'HTTP X-Powered-By', 'HTTP Dir Listing',
        'HTTPS Status', 'HTTPS Final URL', 'HTTPS Server', 'HTTPS X-Powered-By', 'HTTPS Dir Listing',
        'SSL Issuer', 'SSL Expiry Date', 'SSL Days Left', 'SSL Status', 'SSL Error',
        'HSTS (HTTP)', 'X-Frame-Options (HTTP)', 'X-Content-Type-Options (HTTP)',
        'CSP (HTTP)', 'Referrer-Policy (HTTP)', 'Permissions-Policy (HTTP)',
        'HSTS (HTTPS)', 'X-Frame-Options (HTTPS)', 'X-Content-Type-Options (HTTPS)',
        'CSP (HTTPS)', 'Referrer-Policy (HTTPS)', 'Permissions-Policy (HTTPS)',
        'Potential Issues', 'JS Secrets/Errors Found' # New Column
    ]
    sheet.append(headers)

    for item in data_list:
        # Ensure all js_secrets_found items are strings
        js_secrets_str_list = [str(s) for s in item.get('js_secrets_found', ['N/A'])]
        
        row = [
            item.get('target', 'N/A'),
            item.get('ip_resolved_for_scan', 'N/A'),
            ", ".join(item.get('A', [])), ", ".join(item.get('AAAA', [])),
            ", ".join(item.get('CNAME', [])), ", ".join(item.get('MX', [])),
            ", ".join(item.get('NS', [])), "; ".join(item.get('TXT', []) if isinstance(item.get('TXT'), list) else [str(item.get('TXT', ''))]),
            item.get('open_ports', 'N/A'),
            item.get('http_status', 'N/A'), item.get('final_url_http', 'N/A'),
            item.get('http_headers', {}).get('Server', 'N/A'), item.get('http_headers', {}).get('X-Powered-By', 'N/A'),
            item.get('directory_listing_http', 'N/A'),
            item.get('https_status', 'N/A'), item.get('final_url_https', 'N/A'),
            item.get('https_headers', {}).get('Server', 'N/A'), item.get('https_headers', {}).get('X-Powered-By', 'N/A'),
            item.get('directory_listing_https', 'N/A'),
            item.get('ssl_issuer', 'N/A'), item.get('ssl_expiry_date', 'N/A'),
            item.get('ssl_days_left', 'N/A'), item.get('ssl_status', 'N/A'), item.get('ssl_error', 'N/A'),
            item.get('security_headers_http', {}).get('Strict-Transport-Security', 'N/A'),
            item.get('security_headers_http', {}).get('X-Frame-Options', 'N/A'),
            item.get('security_headers_http', {}).get('X-Content-Type-Options', 'N/A'),
            item.get('security_headers_http', {}).get('Content-Security-Policy', 'N/A'),
            item.get('security_headers_http', {}).get('Referrer-Policy', 'N/A'),
            item.get('security_headers_http', {}).get('Permissions-Policy', 'N/A'),
            item.get('security_headers_https', {}).get('Strict-Transport-Security', 'N/A'),
            item.get('security_headers_https', {}).get('X-Frame-Options', 'N/A'),
            item.get('security_headers_https', {}).get('X-Content-Type-Options', 'N/A'),
            item.get('security_headers_https', {}).get('Content-Security-Policy', 'N/A'),
            item.get('security_headers_https', {}).get('Referrer-Policy', 'N/A'),
            item.get('security_headers_https', {}).get('Permissions-Policy', 'N/A'),
            item.get('potential_issues', 'N/A'),
            # Join list of found JS secrets/errors with a newline for readability in Excel cell
            "\n".join(js_secrets_str_list) 
        ]
        sheet.append(row)
    
    for col_idx, column_cells in enumerate(sheet.columns):
        max_len = 0
        col_letter = get_column_letter(col_idx + 1)
        for cell in column_cells:
            if cell.value:
                cell_value_str = str(cell.value)
                # For cells with newlines (like JS secrets), consider the longest line
                max_len_cell = max(len(line) for line in cell_value_str.split('\n'))
                if max_len_cell > max_len:
                    max_len = max_len_cell
        # Set a max width to prevent extremely wide columns, e.g., for long TXT records or JS findings
        adjusted_width = min(max_len + 2, 80) # Cap at 80 characters wide
        sheet.column_dimensions[col_letter].width = adjusted_width


    try:
        wb.save(filename)
        print(f"\nResults successfully written to {filename}")
    except Exception as e:
        print(f"\nError writing to Excel file: {e}")
        print("Ensure the file is not open elsewhere.")

def main():
    input_filename = input("Enter the name of the file containing domains/subdomains: ")
    output_filename = input("Enter the desired name for the Excel output file (e.g., results.xlsx): ")
    if not output_filename.endswith(".xlsx"): output_filename += ".xlsx"
    max_threads_str = input("Enter maximum number of concurrent threads (e.g., 10, default: 10): ")
    max_threads = int(max_threads_str) if max_threads_str.isdigit() else 10


    try:
        with open(input_filename, 'r') as f:
            targets = [line.strip() for line in f if line.strip()]
    except FileNotFoundError:
        print(f"Error: Input file '{input_filename}' not found.")
        return
    except Exception as e:
        print(f"Error reading input file: {e}")
        return

    if not targets:
        print("No targets found in the input file.")
        return

    all_results = []
    print(f"\nStarting assessment for {len(targets)} targets using up to {max_threads} threads...\n")

    with ThreadPoolExecutor(max_workers=max_threads) as executor:
        future_to_target = {executor.submit(process_target, target): target for target in targets}
        for i, future in enumerate(as_completed(future_to_target)):
            target = future_to_target[future]
            try:
                result = future.result()
                all_results.append(result)
            except Exception as exc:
                print(f"{target} generated an exception during processing: {exc}")
                all_results.append({'target': target, 'potential_issues': f'Error during processing: {exc}', 'js_secrets_found': [f'Processing Error: {exc}']})
            print(f"Completed: {target} ({i+1}/{len(targets)})")

    if all_results:
        write_to_excel(all_results, output_filename)
    else:
        print("No results were generated.")

if __name__ == "__main__":
    main()
